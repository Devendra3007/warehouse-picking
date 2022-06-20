

from asyncio.windows_events import NULL
from calendar import c
from ntpath import join
import openpyxl


workbook = openpyxl.load_workbook("Book1.xlsx")
ws = workbook["Sheet1"]

number_of_segments = ws.cell(row = 1, column = 1).value
len_racks = ws.cell(row = 2, column = 1).value

    
def getrow(start_col , start_row  ,  start_with , len ,with_rack):

        for i in range(0,len):
            if with_rack : 
                
                cell1 = ws.cell(row = start_row + 1   , column = start_col + i   )
                value = str(start_with) +"-1"
                if cell1.value:
                    values[value] = " ".join(str(cell1.value).split(" ")[0:2])   
                else:
                    if cell1.value is None:
                        values[value] = " ".join(str(cell1.value).split(" ")[0:2])
                    else:    
                        values[value] = " ".join(str(cell1.value).split(" ")[0:2])+"-".join(str(cell1.value).split("-")[0:1])
                    
            start_with +=1
                

def get_column_racks(start_col , start_row , len ,segment_number , y ):
        for i in range(1,len+1):
            circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
            cell = ws.cell(row = start_row + i - 1 , column = start_col )
            value = str(circle_number) + "-" + str(y)
            values[value] = " ".join(str(cell.value).split(" ")[0:2])
            
values = {}

for i in range(0,number_of_segments):   

        getrow(start_col = 4 +i*4 , start_row = 8, start_with = i*3 , len = 3 , with_rack=False)
        get_column_racks(4 + i*4 , 10 , len_racks , i*2 , i*2 + 2)
        get_column_racks(6 + i*4 , 10 , len_racks , i*2 +1 ,  i*2 + 3)
        getrow(start_col = 4 + i*4, start_row = 11 + len_racks  , start_with = i*3 + len_racks*number_of_segments + number_of_segments*3 , len = 3 , with_rack=True)

data = {}
data["number_of_segments"] = number_of_segments
data["number_of_racks"] = len_racks
data["values"] = values

print(data)
