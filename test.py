from asyncio.windows_events import NULL
import csv
import json
from msilib import datasizemask
import MySQLdb
from flask import Flask, redirect, request,session, url_for,flash,render_template
from flask_mysqldb import MySQL
import openpyxl
from flask import jsonify
import pytz
import datetime

app = Flask(__name__,static_url_path='/static')

app.config['MYSQL_HOST'] = "localhost"
app.config['MYSQL_USER'] = "root"
app.config['MYSQL_PASSWORD'] = "toor"
app.config['MYSQL_DB'] = "atlas_copco"
app.secret_key = 'super secret key'
app.config['SESSION_TYPE'] = 'filesystem'
mysql = MySQL(app)

@app.route('/',methods=['GET','POST'])
def homepage():
    return render_template('HomePage.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
   msg = {}
   if request.method == 'POST' and 'email' in request.form and 'password' in request.form:
        
      username = request.form['email']
      password = request.form['password']
        
      cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
      cursor.execute('SELECT * FROM kitter_info WHERE email = %s AND password = %s', (username, password,))
        
      account = cursor.fetchone()
        
      if account:
         msg['kitter_id']=account['kitter_id']
         msg['username'] = account['kitter_name']
         msg["error"] = "101"
          
      else:
         # Account doesnt exist or username/password incorrect
         msg['kitter_id']=NULL
         msg['username'] = NULL
         msg["error"] = "000"
    
   return json.dumps(msg,indent=4)
   

@app.route("/item_list", methods = ['GET'])
def iteam_list():
   #xlsmFile= request.args.get('s_no')
   s= str(request.args.get('s_no'))  
  
   cursor=mysql.connection.cursor()
   
   cursor.execute("select item_details.*, variant_details.quantity_to_picked from item_details,variant_details where item_details.item_id=variant_details.item_id and item_details.item_id IN (SELECT item_id FROM variant_details WHERE variant_no = (SELECT variant_no From serial_no WHERE serial_no = '"+s+"'))")
   data=cursor.fetchall()
   
   ft ={'item_id','item_name','location','shelf_no','total_quanity','barcode','quantiy_to_picked'}
   l=[]

   for i in data:
      ft={'item_id':i[0],'item_name':i[1],'location':i[2],'shelf_no':i[3],'total_quanity':i[4],'barcode':i[5],'quantiy_to_picked':i[6]}
      l.append(ft)

   
   return json.dumps(l,indent=4)

@app.route("/layout", methods = ['GET'])
def get_layout():
    workbook = openpyxl.load_workbook("Book1.xlsx")
    ws = workbook["Sheet1"]

    number_of_segments = ws.cell(row = 1, column = 1).value
    len_racks = ws.cell(row = 2, column = 1).value
    starting_location = ws.cell(row =1 ,column=4).value

    # Program to parse excel file
    def getrow(start_col , start_row  ,  start_with , len ,with_rack):

        for i in range(0,len):
            if with_rack : 
                
                cell1 = ws.cell(row = start_row + 1   , column = start_col + i   )
                value = str(start_with) +"-1"
                if cell1.value:
                    values[value] = " ".join(str(cell1.value).split(" ")[0:2])   
                else:
                    if cell1.value is None:
                        values[value] = " ".join(str(cell1.value).split(" ")[0:2])
                    else:    
                        values[value] = " ".join(str(cell1.value).split(" ")[0:2])+"-".join(str(cell1.value).split("-")[0:1])
                    
            start_with +=1
                

    def get_column_racks(start_col , start_row , len ,segment_number , y ):
        for i in range(1,len+1):
            circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
            cell = ws.cell(row = start_row + i - 1 , column = start_col )
            value = str(circle_number) + "-" + str(y)
            values[value] = " ".join(str(cell.value).split(" ")[0:2])
    values = {}

    for i in range(0,number_of_segments):   

        getrow(start_col = 4 +i*4 , start_row = 8, start_with = i*3 , len = 3 , with_rack=False)
        get_column_racks(4 + i*4 , 10 , len_racks , i*2 , i*2 + 2)
        get_column_racks(6 + i*4 , 10 , len_racks , i*2 +1 ,  i*2 + 3)
        getrow(start_col = 4 + i*4, start_row = 11 + len_racks  , start_with = i*3 + len_racks*number_of_segments + number_of_segments*3 , len = 3 , with_rack=True)

    data = {}
    data["number_of_segments"] = number_of_segments
    data["number_of_racks"] = len_racks
    data["staring_location"] = starting_location
    data["values"] = values
    
    return json.dumps(data , indent = 4)

@app.route("/items_picked", methods =['POST'])
def items_picked():
    d_month = datetime.datetime.now(tz=pytz.timezone('Asia/Kolkata'))
    now = datetime.datetime.now()
    current_time = now.strftime("%H_%M_%S")
    date=d_month.strftime('%B_%d_%Y')
    date=date.lower()

    if request.method == 'POST':
        data = request.get_json()
        csvfilename=data[0]["kitter_name"]+"_"+date+"_"+current_time+ "_done" + ".csv"
        
        fields = ("kitter_id","kitter_name","item_id","item_name","quantity_picked")
        with open (csvfilename,"w") as file:
            writer = csv.DictWriter(file, fieldnames=fields)
            writer.writeheader()
            for i in data:
                
                cur =mysql.connection.cursor()
                cur.execute("UPDATE item_details SET total_quantity=total_quantity-"+str(i.get("quantity_picked"))+" WHERE item_id = '"+i.get("item_id")+"'")
                mysql.connection.commit()
                cur.close()
                writer.writerow(i)
        return json.dumps("done")



#delete data 


@app.route('/Add_Serial_no',methods=['GET','POST'])
def Add_Serial_info():
    if request.method == "POST":
        data = request.form
        Serial_no = data['Serial_no']
        Varient_no = data['Varient_no']
        cur =mysql.connection.cursor()

        x = cur.execute("SELECT * FROM serial_no WHERE serial_no ='"+Serial_no+"'")
        if int(x) > 0:
           flash("Serial No is already taken")
           
        else:
            cur.execute("INSERT INTO serial_no VALUES(%s,%s)",(Serial_no,Varient_no))
            mysql.connection.commit()
            cur.close()
            flash('Data Saved!!')
        
    return render_template('Add_Data_Serial_No.html')

@app.route('/Delete_Serial_no',methods=['GET','POST'])
def Delete_Serial_info():
    if request.method == "POST":
        data = request.form
        Serial_no = data['Serial_no']
        cur =mysql.connection.cursor()

        x = cur.execute("SELECT * FROM serial_no WHERE serial_no ='"+Serial_no+"'")
        if int(x) == 0:
           flash("Serial Number Dose not exist")
        else:
            cur.execute("DELETE FROM serial_no WHERE serial_no ='"+Serial_no+"'")
            mysql.connection.commit()
            cur.close()
            flash('Data Delete!!')
        
    return render_template('Delete_Serial_No.html')

@app.route('/Add_Variant_Details',methods=['GET','POST'])
def Add_Variant_Details():
    if request.method == "POST":
        data = request.form
        Item_Id = data['Item_Id']
        Varient_no = data['Varient_no']
        Required_Quantity = data['Required_Quantity']
        cur =mysql.connection.cursor()

        x = cur.execute("SELECT * FROM variant_details WHERE item_id ='"+Item_Id+"'")
        if int(x) > 0:
           flash("That name is already taken, please choose another")
           
        else:
            cur.execute("INSERT INTO variant_details VALUES(%s,%s,%s)",(Varient_no,Item_Id,Required_Quantity))
            mysql.connection.commit()
            cur.close()
            flash('Data Saved!!')
        
    return render_template('Add_Variant_Details.html') 

@app.route('/Delete_Variant_Details', methods=['GET', 'POST'])
def Delete_Variant_Details():
    if request.method == "POST":
        details2 = request.form
        
        Variant_no = details2['Variant_no']
        Item_Id = details2['Item_Id']
        
        cur = mysql.connection.cursor()
        x =cur.execute("SELECT * FROM variant_details WHERE variant_no=%s and item_id=%s ",(Variant_no,Item_Id))
        if int(x) ==0:
            flash("Variant Number Not Present")

        else:
            cur.execute("DELETE FROM variant_details where variant_no=%s and item_id=%s ",(Variant_no,Item_Id))
            mysql.connection.commit()
            cur.close()
            flash("Deleted Variant")
    return render_template('Delete_Variant_Details.html')

@app.route('/Add_Item_Details',methods=['GET','POST'])
def Add_Item_Details():
    if request.method == "POST":
        data = request.form
        Item_Id = data['Item_Id']
        Item_Name = data['Item_Name']
        Item_Location = data['Item_Location']
        Barecode = data['Barecode']
        Self =data['Self']
        Total_Quantity = data['Total_Quantity']
        cur =mysql.connection.cursor()

        x = cur.execute("SELECT * FROM item_details WHERE item_id ='"+Item_Id+"'")
        if int(x) > 0:
           flash("That Item_id is already taken, please choose another")
           
        else:
            cur.execute("INSERT INTO item_details VALUES(%s,%s,%s,%s,%s,%s)",(Item_Id,Item_Name
                        ,Item_Location,Self,Total_Quantity,Barecode))
            mysql.connection.commit()
            cur.close()
            flash('Data Saved!!')
        
    return render_template('Add_Item_Details.html')

@app.route('/Add_Item_Total_Quantity',methods=['GET','POST'])
def Add_Item_Total_Quantity():
    if request.method == "POST":
        data = request.form
        Item_Id = data['Item_Id']
        Add_Quantity = data['Add_Quantity']
        cur =mysql.connection.cursor()

        x= cur.execute("SELECT * FROM item_details WHERE item_id ='"+Item_Id+"'")
        if int(x) == 0:
            flash("Item Related To Item id =%s is not avaliable",(Item_Id))
        else:
            cur.execute("UPDATE item_details SET total_quantity=total_quantity+"+Add_Quantity+" WHERE item_id = '"+Item_Id+"'")
            mysql.connection.commit()
            cur.close()
            flash('Quntity Updated!!')
    return render_template('Add_Item_Total_Quantity.html')

@app.route('/Edit_Item_Detail',methods=['GET','POST'])
def Edit_Item_Details():
    if request.method == "POST":
        data = request.form
        Item_Id = data['Item_Id']
        Item_Name = data['Item_Name']
        Item_Location = data['Item_Location']
        Shelf =data['Self']
        Total_Quantity = data['Total_Quantity']
        cur =mysql.connection.cursor()
        x= cur.execute("SELECT * FROM item_details WHERE item_id ='"+Item_Id+"'")
        if int(x) == 0:
            flash("Item Related To Item_Id is not avaliable")
        else:
            cur.execute("UPDATE item_details SET item_name= '"+Item_Name+"' WHERE item_id = '"+Item_Id+"'")
            cur.execute("UPDATE item_details SET location = '"+Item_Location+"' WHERE item_id = '"+Item_Id+"'")
            if int(Shelf)>0:
                cur.execute("UPDATE item_details SET shelf_no = '"+Shelf+"' WHERE item_id = '"+Item_Id+"'")
            if int(Total_Quantity)>=0:
                cur.execute("UPDATE item_details SET total_quantity = '"+Total_Quantity+"' WHERE item_id = '"+Item_Id+"'")
            mysql.connection.commit()
            cur.close()
            flash('Updated!!')
    return render_template('Edit_Item_Detail.html')

@app.route('/Delete_Item_Details', methods=['GET', 'POST'])
def Delete_Item_Details():
    if request.method == "POST":
        details2 = request.form
        
        Item_Id = details2['Item_Id']
        cur = mysql.connection.cursor()
        x =cur.execute("SELECT * FROM item_details WHERE  item_id=%s ",[Item_Id])
        if int(x) ==0:
            flash("Item Id Not Present")

        else:
            cur.execute("DELETE FROM item_details where  item_id=%s ",[Item_Id])
            mysql.connection.commit()
            cur.close()
            flash("Item Deleted")
    return render_template('Delete_Item_Details.html')

@app.route("/current_location", methods = ['GET'])
def current_location():
   
   s= str(request.args.get('current_location')) 
   

   cursor=mysql.connection.cursor()
   
   cursor.execute("select location From item_details where item_id = %s",[s])
   locations=cursor.fetchone()
   
   data ={}
   data["locaiton"] = locations
   
   return json.dumps(data,indent=4)
    
if __name__ == '__main__':
   app.run(debug=True,host="https://warehousepicking.herokuapp.com") 